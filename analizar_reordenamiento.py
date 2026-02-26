import openpyxl
import pandas as pd

# Analizar archivo de Reordenamiento
wb = openpyxl.load_workbook('Reordenamiento XLSX (2).xlsx')
print("Hojas disponibles:", wb.sheetnames)
ws = wb.active

print("\n" + "="*80)
print("PRIMERAS 15 FILAS DEL ARCHIVO")
print("="*80)

for i in range(1, 16):
    row = [ws.cell(row=i, column=j).value for j in range(1, 15)]
    print(f"Fila {i}: {row}")

print("\n" + "="*80)
print("ANÁLISIS CON PANDAS")
print("="*80)

try:
    df = pd.read_excel('Reordenamiento XLSX (2).xlsx', sheet_name=0)
    print(f"\nForma del dataframe: {df.shape}")
    print(f"\nColumnas:")
    for i, col in enumerate(df.columns):
        print(f"  {i}: {col}")
    
    print(f"\nPrimeras 10 filas:")
    print(df.head(10))
    
    print(f"\nValores únicos en columna 0:")
    print(df.iloc[:, 0].unique()[:10])
except Exception as e:
    print(f"Error: {e}")
